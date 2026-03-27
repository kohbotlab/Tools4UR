"""
Excel Test Report Generator

Generates Excel report for test result.
Requires openpyxl (pip install openpyxl). Graceful skip if not installed.

Shared module used by send_to_ursim.py and send_to_robot.py.
run_test() returns data only; caller calls save_report().
Do NOT call run_test() from here — this module only formats data.
"""

import os
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    HAS_OPENPYXL = True

    _FONT_NAME = "Meiryo UI"
    _THIN_BORDER = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    _HEADER_FILL = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid")
    _HEADER_FONT = Font(
        name=_FONT_NAME, bold=True, size=10, color="FFFFFF")
    _TITLE_FONT = Font(name=_FONT_NAME, bold=True, size=12)
    _LABEL_FONT = Font(name=_FONT_NAME, bold=True, size=10)
    _BASE_FONT = Font(name=_FONT_NAME, size=10)
    _PASS_FILL = PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    _FAIL_FILL = PatternFill(
        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    _PASS_FONT = Font(
        name=_FONT_NAME, bold=True, size=10, color="006100")
    _FAIL_FONT = Font(
        name=_FONT_NAME, bold=True, size=10, color="9C0006")
    _SECTION_FONT = Font(
        name=_FONT_NAME, bold=True, size=11, color="4472C4")

except ImportError:
    HAS_OPENPYXL = False


# ---- Helpers ----

def _cell(ws, row, col, value, font=None, fill=None, alignment=None,
          border=True):
    """Write a cell with optional styling."""
    c = ws.cell(row=row, column=col, value=value)
    c.font = font or _BASE_FONT
    if fill:
        c.fill = fill
    if alignment:
        c.alignment = alignment
    if border:
        c.border = _THIN_BORDER
    return c


def _header_row(ws, row, headers, start_col=1):
    """Write a styled header row."""
    for col, h in enumerate(headers, start=start_col):
        _cell(ws, row, col, h, font=_HEADER_FONT, fill=_HEADER_FILL,
              alignment=Alignment(horizontal="center"))


# ---- Sheet writer ----

def _write_test_sheet(wb, result, log_filename=None,
                      sensor_log_filename=None):
    """Write one sheet for a test result (two-panel layout).

    Layout:
      Title:         B2, merged B2:J2
      Left panel:    B-C, rows 3-17 (summary fields)
      Right panel:
        Verification:    E-H, rows 3-8
        Cycle Results:   E-N, rows 11+
        Fixed Registers: E-G, dynamic rows after cycles
    """
    ws = wb.active
    ws.title = "Result"

    # --- Column widths ---
    ws.column_dimensions["A"].width = 2.9
    ws.column_dimensions["B"].width = 17.5
    ws.column_dimensions["C"].width = 39.7
    ws.column_dimensions["D"].width = 9.6
    ws.column_dimensions["E"].width = 18
    for letter in ["F", "G", "H", "I", "J", "K"]:
        ws.column_dimensions[letter].width = 13
    ws.column_dimensions["L"].width = 12.6
    for letter in ["M", "N"]:
        ws.column_dimensions[letter].width = 13

    # --- Title at B2, merged B2:J2 ---
    desc = result.get("description", "")
    _cell(ws, 2, 2, desc, font=_TITLE_FONT, border=False)
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=10)

    # --- Left panel (B-C, rows 3-15): summary fields ---
    verdict = result.get("verdict", "?")
    is_pass = verdict == "PASS"

    summary_pairs = [
        ("Target", result.get("target", "")),
        ("Timestamp", result.get("timestamp", "")),
        ("Script File(s)", ", ".join(result.get("script_files", []))),
        ("Param Verify",
         f"{result.get('verify_pass', 0)} OK / "
         f"{result.get('verify_fail', 0)} NG"),
        ("Cycle Count",
         f"{result.get('cycle_count_initial', '?')} -> "
         f"{result.get('cycle_count_final', '?')} "
         f"(delta={result.get('cycle_count_delta', '?')})"),
        ("Verdict", verdict),
        ("Cycles",
         f"{result.get('completed', 0)}/{result.get('total', 0)}"),
        ("Avg Time", f"{result.get('avg_time', 0):.1f}s"),
        ("Send Error",
         "No" if not result.get("send_error", False) else "Yes"),
        ("Picks", result.get("picks", "")),
        ("Log File", log_filename or ""),
    ]
    if sensor_log_filename:
        summary_pairs.append(("Sensor Log", sensor_log_filename))

    for i, (label, value) in enumerate(summary_pairs):
        row = 3 + i
        _cell(ws, row, 2, label, font=_LABEL_FONT)
        vc = _cell(ws, row, 3, value)
        if label == "Verdict":
            vc.fill = _PASS_FILL if is_pass else _FAIL_FILL
            vc.font = _PASS_FONT if is_pass else _FAIL_FONT

    # --- Right panel: Verification (E-H, rows 3-8) ---
    _cell(ws, 3, 5, "Verification", font=_SECTION_FONT, border=False)
    ws.merge_cells(start_row=3, start_column=5, end_row=3, end_column=7)

    v_headers = ["Check", "Result", "Detail"]
    _header_row(ws, 4, v_headers, start_col=5)

    checks = [
        ("Param Verify",
         result.get("verify_fail", 0) == 0,
         f"{result.get('verify_pass', 0)} OK / "
         f"{result.get('verify_fail', 0)} NG",
         None),
        ("Cycle Count",
         result.get("cycle_count_ok", True),
         f"{result.get('cycle_count_initial', '?')} -> "
         f"{result.get('cycle_count_final', '?')}",
         f"(delta={result.get('cycle_count_delta', '?')})"),
        ("Picks (cycle_ready)",
         result.get("pick_ok", True),
         f"{result.get('picks', '?')}/{result.get('completed', '?')}",
         None),
        ("Send Error",
         not result.get("send_error", False),
         "No" if not result.get("send_error", False) else "Yes",
         None),
    ]
    for i, (name, passed, detail, extra) in enumerate(checks):
        row = 5 + i
        _cell(ws, row, 5, name)
        rc = _cell(ws, row, 6, "OK" if passed else "NG")
        rc.fill = _PASS_FILL if passed else _FAIL_FILL
        rc.font = _PASS_FONT if passed else _FAIL_FONT
        _cell(ws, row, 7, detail)
        if extra:
            _cell(ws, row, 8, extra)

    # --- Right panel: Cycle Results (E-N, rows 11+) ---
    _cell(ws, 11, 5, "Cycle Results", font=_SECTION_FONT, border=False)
    ws.merge_cells(start_row=11, start_column=5, end_row=11, end_column=12)

    cycle_headers = [
        "#", "offset_x", "offset_y",
        "target_pallet", "layer_height",
        "Time (s)", "Verify", "Mismatches",
    ]
    _header_row(ws, 12, cycle_headers, start_col=5)

    cycle_details = result.get("cycle_details", [])
    for r_offset, cd in enumerate(cycle_details):
        r = 13 + r_offset
        params = cd.get("params", {})
        values = [
            cd.get("cycle_num", r_offset + 1),
            params.get("offset_x", ""),
            params.get("offset_y", ""),
            params.get("target_pallet", ""),
            params.get("layer_height", ""),
            f"{cd.get('time', 0):.1f}",
            "OK" if cd.get("verify") else "NG",
            ", ".join(cd.get("mismatches", [])),
        ]
        for col_offset, val in enumerate(values):
            _cell(ws, r, 5 + col_offset, val)

        # Color verify cell (column 5 + 6 = 11)
        v_cell = ws.cell(row=r, column=11)
        if cd.get("verify"):
            v_cell.fill = _PASS_FILL
            v_cell.font = _PASS_FONT
        else:
            v_cell.fill = _FAIL_FILL
            v_cell.font = _FAIL_FONT

    # --- Right panel: Fixed Registers (E-G, after cycles) ---
    fixed_regs = result.get("fixed_registers", {})
    if fixed_regs:
        cycle_end_row = 13 + len(cycle_details)
        fr_header_row = cycle_end_row + 2  # 2 blank rows

        _cell(ws, fr_header_row, 5, "Fixed Registers",
              font=_SECTION_FONT, border=False)
        ws.merge_cells(start_row=fr_header_row, start_column=5,
                       end_row=fr_header_row, end_column=7)

        fr_col_row = fr_header_row + 1
        fr_headers = ["Register", "Name", "Value"]
        for col_offset, h in enumerate(fr_headers):
            _cell(ws, fr_col_row, 5 + col_offset, h,
                  font=_HEADER_FONT, fill=_HEADER_FILL,
                  alignment=Alignment(horizontal="left"))

        data_row = fr_col_row + 1
        for addr, (name, value) in sorted(fixed_regs.items()):
            _cell(ws, data_row, 5, addr)
            _cell(ws, data_row, 6, name)
            _cell(ws, data_row, 7, value)
            data_row += 1


# ---- Public API ----

def save_report(result, report_dir=None, session_ts=None, log_file=None,
                sensor_log_file=None):
    """Generate Excel report for a test result.

    Args:
        result: result_dict from run_test().
        report_dir: directory to save (default: report/ next to this file)
        session_ts: timestamp string (YYYYMMDD_HHMMSS) to match log filename.
            If None, generates a new timestamp.
        log_file: path to the main log file for this session (shown in report).
        sensor_log_file: path to the sensor log file (URSim only, shown in report).

    Returns:
        filepath (str) on success, None on skip/error.
    """
    if not HAS_OPENPYXL:
        print("  [Report] openpyxl not installed, skipping report generation")
        print("  [Report] Install: pip install openpyxl")
        return None

    if not result or result.get("verdict") is None:
        return None

    if report_dir is None:
        report_dir = os.path.join(
            os.path.dirname(os.path.abspath(__file__)), "report")
    os.makedirs(report_dir, exist_ok=True)

    # File naming
    target = result.get("target", "unknown").replace(" ", "")
    ts = session_ts or datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"Report_{target}_{ts}.xlsx"
    filepath = os.path.join(report_dir, filename)

    try:
        wb = Workbook()
        log_filename = os.path.basename(log_file) if log_file else None
        sensor_log_filename = (os.path.basename(sensor_log_file)
                               if sensor_log_file else None)
        _write_test_sheet(wb, result,
                          log_filename=log_filename,
                          sensor_log_filename=sensor_log_filename)
        wb.save(filepath)
        print(f"  [Report] Saved: {filepath}")
        return filepath
    except Exception as e:
        print(f"  [Report] Error: {e}")
        return None
